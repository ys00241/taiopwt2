#!/usr/bin/env python3
"""
Import 2025 sorting data (sorting2025.xlsx) into taio_pwt DB.
Creates items, this_year_items, members, and bids from the Excel.
Run inside the Docker container:
    docker cp sorting2025.xlsx taio_pwt:/app/
    docker exec taio_pwt python scripts/import_2025.py
"""
import sys
import os
import uuid
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

os.environ.setdefault("FLASK_ENV", "production")

from app import create_app
from app.extensions import db
from app.models.member import Member
from app.models.item import Item
from app.models.this_year_item import ThisYearItem
from app.models.bid import Bid
from app.models.edition import Edition

import openpyxl


def get_or_create_member(name):
    """Find member by name, or create a stub member."""
    if not name or str(name).strip() == "":
        return None
    name = str(name).strip()
    member = Member.query.filter_by(name=name).first()
    if member:
        return member
    # Get max member_id
    max_id = db.session.query(db.func.max(Member.member_id)).scalar() or 99
    new_id = max_id + 1
    member = Member(
        id=str(uuid.uuid4()),
        member_id=new_id,
        name=name,
        member_type="member",
        status="active",
    )
    db.session.add(member)
    db.session.flush()
    print(f"  👤 Created new member: {name} (#{new_id})")
    return member


def get_or_create_item(name_1, name_2, category, cost, supplier):
    """Find item by name_1_auspicious, or create new item master record."""
    if not name_1 or str(name_1).strip() == "":
        name_1 = name_2 or "未命名"
    name_1 = str(name_1).strip()
    name_2 = str(name_2).strip() if name_2 else ""
    cost_val = float(cost) if cost else 0
    supplier_val = str(supplier).strip() if supplier else ""

    # Try to find existing item by name_1_auspicious
    item = Item.query.filter_by(name_1_auspicious=name_1).first()
    if not item and name_2:
        item = Item.query.filter_by(name_2_description=name_2).first()
    if item:
        return item

    # Get max item_id
    max_item_id = db.session.query(db.func.max(Item.item_id)).scalar() or 0
    new_item_id = max_item_id + 1

    # Store year_data as JSON
    import json
    year_data = json.dumps([{
        "year": 2025,
        "cost_hkd": cost_val,
        "supplier": supplier_val,
    }], ensure_ascii=False)

    item = Item(
        item_id=new_item_id,
        name_1_auspicious=name_1,
        name_2_description=name_2 or None,
        year_data=year_data,
    )
    db.session.add(item)
    db.session.flush()
    print(f"  📦 Created new item: {name_1} (ID#{new_item_id})")
    return item


def main():
    # Find Excel file
    excel_path = Path("/app/sorting2025.xlsx")
    if not excel_path.exists():
        # Also check /tmp
        excel_path = Path("/tmp/sorting2025.xlsx")
    if not excel_path.exists():
        print("❌ sorting2025.xlsx not found! Copy it to container first:")
        print("   docker cp sorting2025.xlsx taio_pwt:/app/")
        return

    app = create_app()
    with app.app_context():
        db.create_all()

        print("=" * 60)
        print("📥 Importing 2025 sorting data...")
        print("=" * 60)

        wb = openpyxl.load_workbook(str(excel_path), data_only=True)
        ws = wb["今年聖物2026"]

        bid_no = 0
        items_created = 0
        members_created = 0
        this_year_items_created = 0
        bids_created = 0

        for r in range(2, ws.max_row + 1):
            year_val = ws.cell(r, 1).value
            sticker_no = ws.cell(r, 2).value
            name_1 = ws.cell(r, 3).value  # 聖物名稱
            name_2 = ws.cell(r, 4).value  # 實際物品
            category = ws.cell(r, 5).value or "其他"
            cost = ws.cell(r, 6).value or 0
            supplier = ws.cell(r, 7).value or ""
            bidder = ws.cell(r, 8).value  # 投得者
            bid_amount = ws.cell(r, 9).value or 0
            handler = ws.cell(r, 10).value or ""
            photo_ref = ws.cell(r, 11).value or ""
            remarks = ws.cell(r, 12).value or ""

            if not name_1 and not name_2:
                continue  # skip empty rows

            # Normalize sticker number
            try:
                sticker_no = int(float(str(sticker_no).strip()))
            except (ValueError, TypeError):
                sticker_no = 0

            # Get or create item master
            item = get_or_create_item(name_1, name_2, category, cost, supplier)
            if not item:
                continue

            # Get or create member
            member = get_or_create_member(bidder) if bidder else None

            # Create ThisYearItem
            existing_tyi = ThisYearItem.query.filter_by(
                year=year_val, sticker_no=sticker_no
            ).first()
            if not existing_tyi:
                tyi = ThisYearItem(
                    year=year_val or 2025,
                    sticker_no=sticker_no,
                    item_id=item.item_id,
                    cost=float(cost) if cost else 0,
                    supplier=str(supplier).strip() if supplier else "",
                    photo_ref=str(photo_ref).strip() if photo_ref else "",
                    remarks=str(remarks).strip() if remarks else "",
                )
                db.session.add(tyi)
                this_year_items_created += 1

            # Create Bid (only if there's a bidder and amount > 0)
            if member and bid_amount and float(bid_amount) > 0:
                bid_no = (bid_no % 999) + 1
                bid = Bid(
                    id=str(uuid.uuid4()),
                    year=year_val or 2025,
                    bid_no=bid_no,
                    item_id=item.item_id,
                    member_id=member.member_id,
                    bid_amount=float(bid_amount),
                    paid_amount=0,
                    handler=str(handler).strip() if handler else "",
                    payment_method="cash",
                )
                db.session.add(bid)
                bids_created += 1

            items_created += 1

        # Ensure Edition 2025 exists
        if not Edition.query.filter_by(year=2025).first():
            ed = Edition(
                id=str(uuid.uuid4()),
                year=2025,
                edition_no=1,
                event_date="2025-03-01",
                status="completed",
            )
            db.session.add(ed)

        db.session.commit()
        wb.close()

        print("\n" + "=" * 60)
        print("✅ Import complete!")
        print(f"   Items processed: {items_created}")
        print(f"   New members created: {Member.query.count()}")
        print(f"   ThisYearItems: {this_year_items_created}")
        print(f"   Bids created: {bids_created}")
        print("=" * 60)


if __name__ == "__main__":
    main()
