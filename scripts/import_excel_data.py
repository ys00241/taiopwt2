#!/usr/bin/env python3
"""
Comprehensive Excel data import script for 寶榮堂花炮會 (taiopwt2).

Reads input_data.xlsx and imports members, items, this_year_items, and bids
for years 2015–2026 into the local SQLite database.

Usage:
    source .venv/bin/activate
    python scripts/import_excel_data.py

Output:
    Prints import summary (counts per table, per year).
"""

import json
import os
import sys
from collections import OrderedDict

# ── Flask app bootstrap ──────────────────────────────────────────────────
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
os.environ.setdefault("FLASK_ENV", "development")

from app import create_app
from app.extensions import db
from app.models import Member, Item, ThisYearItem, Bid
# sqlalchemy is available via Flask-SQLAlchemy

import openpyxl

# ── Paths ────────────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH = os.path.join(BASE_DIR, "uploads", "input_data.xlsx")
DB_PATH = os.path.join(BASE_DIR, "data", "paopao.db")

# ── Per-year sheet configuration ─────────────────────────────────────────
# Each entry defines:
#   sheet       : actual sheet name in the workbook
#   header_row  : 1-indexed row containing column headers
#   data_start  : 1-indexed first data row
#   columns     : mapping of logical field -> 1-indexed column number
# The column mappings were determined by inspecting each sheet's header row.

YEAR_LAYOUTS = [
    {
        "year": 2026,
        "sheet": "2026",
        "header_row": 4,
        "data_start": 5,
        "columns": {
            "sticker_no": 2,
            "item_name": 3,
            "actual_item": 4,
            "bidder_name": 5,
            "bid_amount": 6,
            "fee": 7,
            "paid": 9,
            "handler": 10,
            "photo_codes": 11,
            "cost": 12,
            "source": 13,
        },
    },
    {
        "year": 2025,
        "sheet": "2025",
        "header_row": 4,
        "data_start": 5,
        "columns": {
            "sticker_no": 2,
            "item_name": 3,
            "actual_item": 4,
            "bidder_name": 5,
            "bid_amount": 6,
            "fee": 7,
            "paid": 9,
            "handler": 10,
            "photo_codes": 11,
            "cost": 12,
            "source": 13,
        },
    },
    {
        "year": 2024,
        "sheet": "2024",
        "header_row": 4,
        "data_start": 5,
        "columns": {
            "sticker_no": 2,
            "item_name": 3,
            "actual_item": 4,
            "bidder_name": 5,
            "bid_amount": 6,
            "fee": 7,
            "paid": 9,
            "handler": 10,
            "photo_codes": 11,
            "cost": 12,
            "source": 13,
        },
    },
    {
        "year": 2023,
        "sheet": "2023",
        "header_row": 4,
        "data_start": 5,
        "columns": {
            "sticker_no": 2,
            "item_name": 3,
            "actual_item": 4,
            "bidder_name": 5,
            "bid_amount": 6,
            "fee": 7,
            "paid": 9,
            "handler": 10,
            "photo_codes": 11,
            "cost": 12,
            "source": 13,
        },
    },
    {
        "year": 2018,
        "sheet": "2018",
        "header_row": 4,
        "data_start": 5,
        "columns": {
            "sticker_no": 2,
            "item_name": 3,
            "actual_item": 4,
            "bidder_name": 5,
            "bid_amount": 6,
            "fee": 7,
            "paid": 9,
            "handler": 11,  # 經手人 at col 11 for 2018 (col 10 = 抽獎卷)
            "photo_codes": 12,
            "cost": 13,
            # no source column in 2018
        },
    },
    {
        "year": 2017,
        "sheet": " 2017",  # note: sheet name has leading space
        "header_row": 4,
        "data_start": 5,
        "columns": {
            "sticker_no": 2,
            "bidder_name": 3,  # 姓名 is col 3 in 2017
            "item_name": 4,    # 投物名稱1 is col 4
            "actual_item": 5,  # 投物名稱2 is col 5
            "bid_amount": 6,
            "fee": 7,
            "paid": 9,
            "handler": 12,
            "photo_codes": 13,
        },
    },
    {
        "year": 2016,
        "sheet": " 2016",  # note: sheet name has leading space
        "header_row": 3,
        "data_start": 4,
        "columns": {
            "sticker_no": 2,
            "bidder_name": 3,  # 姓名 is col 3
            "item_name": 4,    # 投物名稱1 is col 4
            "actual_item": 5,  # 投物名稱2 is col 5
            "bid_amount": 6,
            "fee": 7,
            "paid": 9,
            "handler": 11,
            "photo_codes": 12,
        },
    },
    {
        "year": 2015,
        "sheet": "2015",
        "header_row": 3,
        "data_start": 4,
        "columns": {
            # 2015 has NO (item number) in col A, no separate sticker_no
            "sticker_no": 1,   # using col 1 as item number
            "bidder_name": 2,  # 姓名 is col 2
            "item_name": 3,    # 投物名稱1 is col 3
            "actual_item": 4,  # 投物名稱2 is col 4
            "bid_amount": 5,
            "fee": 6,
            "paid": 8,
            "handler": 10,
            "photo_codes": 11,
        },
    },
]


# ── Helper functions ─────────────────────────────────────────────────────

def safe_float(val):
    """Convert a cell value to float, returning None if not possible."""
    if val is None:
        return None
    if isinstance(val, str):
        val = val.strip()
        if not val or val in ("", "-"):
            return None
        # Remove currency symbols and commas
        val = val.replace("HK$", "").replace("$", "").replace(",", "").strip()
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def safe_int(val):
    """Convert a cell value to int, returning None if not possible."""
    f = safe_float(val)
    if f is None:
        return None
    return int(f)


def safe_str(val):
    """Convert a cell value to stripped string, returning empty string for None."""
    if val is None:
        return ""
    s = str(val).strip()
    # Clean up newlines in cell text
    s = s.replace("\\n", " ").replace("\n", " ").replace("\r", " ")
    # Collapse multiple spaces
    parts = s.split()
    return " ".join(parts) if parts else ""


def is_total_row(bidder_name):
    """Check if a row is a summary/Total footer row."""
    if bidder_name is None:
        return False
    if isinstance(bidder_name, str):
        return bidder_name.strip().lower() == "total"
    return False


def has_data(row_values, columns_dict):
    """Check if a data row has at least one non-None, non-zero value in mapped columns."""
    for field, col in columns_dict.items():
        val = row_values.get(col)
        if val is not None:
            if isinstance(val, (int, float)) and val != 0:
                return True
            if isinstance(val, str) and val.strip():
                return True
    return False


# ── Main import routine ──────────────────────────────────────────────────

def main():
    app = create_app()
    with app.app_context():
        print(f"Database: {app.config.get('SQLALCHEMY_DATABASE_URI')}")
        print(f"Excel file: {EXCEL_PATH}")
        print()

        # ── 1. Create all tables ──
        db.create_all()
        print("✓ Tables created / verified.")

        # ── 2. Clear existing data in dependency order ──
        Bid.query.delete()
        ThisYearItem.query.delete()
        Item.query.delete()
        Member.query.delete()
        db.session.commit()
        print("✓ Existing data cleared.")

        # ── 3. Trackers ──
        member_registry = OrderedDict()  # name -> member_id
        item_registry = OrderedDict()    # (item_key) -> item_id
        next_member_id = 1
        next_item_id = 1

        # Statistics per year
        stats = {}  # year -> {members, items, this_year_items, bids}

        # Load workbook
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        print()

        # ── 4. First pass: collect all unique member names and item names ──
        print("─── PASS 1: Scanning for unique members and items ───")
        for layout in YEAR_LAYOUTS:
            year = layout["year"]
            sheet_name = layout["sheet"]
            cols = layout["columns"]

            if sheet_name not in wb.sheetnames:
                print(f"  ⚠ Sheet '{sheet_name}' not found, skipping year {year}")
                continue

            ws = wb[sheet_name]
            print(f"  Scanning {sheet_name} (year {year})...")

            for row_idx in range(layout["data_start"], ws.max_row + 1):
                row_vals = {c: ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)}

                # Extract bidder_name
                bidder_name_col = cols.get("bidder_name")
                bidder_name = safe_str(row_vals.get(bidder_name_col)) if bidder_name_col else ""

                # Skip Total rows
                if is_total_row(row_vals.get(bidder_name_col)):
                    continue

                # Skip empty rows (no meaningful data)
                if not has_data(row_vals, cols):
                    continue

                # Collect member (bidder)
                if bidder_name and bidder_name not in member_registry:
                    member_registry[bidder_name] = next_member_id
                    next_member_id += 1

                # Collect item key
                item_name = safe_str(row_vals.get(cols.get("item_name"))) if cols.get("item_name") else ""
                actual_item = safe_str(row_vals.get(cols.get("actual_item"))) if cols.get("actual_item") else ""
                cost = safe_float(row_vals.get(cols.get("cost"))) if cols.get("cost") else 0
                source = safe_str(row_vals.get(cols.get("source"))) if cols.get("source") else ""

                item_key = (item_name, actual_item)
                if item_key not in item_registry and (item_name or actual_item):
                    item_registry[item_key] = next_item_id
                    next_item_id += 1

        print(f"  Found {len(member_registry)} unique members, {len(item_registry)} unique items.")
        print()

        # ── 5. Create Member records ──
        print("─── Creating Member records ───")
        members_created = 0
        for name, mid in member_registry.items():
            member = Member(
                member_id=mid,
                name=name,
                phone="",
            )
            db.session.add(member)
            members_created += 1
        db.session.commit()
        print(f"  ✓ Created {members_created} Member records.")
        print()

        # ── 6. Create Item records (master list) ──
        print("─── Creating Item records (master list) ───")
        items_created = 0
        for (item_name, actual_item), iid in item_registry.items():
            item = Item(
                item_id=iid,
                name_1_auspicious=item_name,
                name_2_description=actual_item,
                year_data=json.dumps([]),  # Will be populated per-year
            )
            db.session.add(item)
            items_created += 1
        db.session.commit()
        print(f"  ✓ Created {items_created} Item records.")
        print()

        # ── 7. Second pass: create ThisYearItem and Bid records per year ──
        print("─── PASS 2: Importing per-year items and bids ───")

        total_this_year_items = 0
        total_bids = 0

        for layout in YEAR_LAYOUTS:
            year = layout["year"]
            sheet_name = layout["sheet"]
            cols = layout["columns"]

            if sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]
            year_stats = {"items": 0, "bids": 0}

            for row_idx in range(layout["data_start"], ws.max_row + 1):
                row_vals = {}
                for c in range(1, ws.max_column + 1):
                    row_vals[c] = ws.cell(row=row_idx, column=c).value

                # ── Extract fields using column map ──
                item_name = safe_str(row_vals.get(cols.get("item_name"))) if cols.get("item_name") else ""
                actual_item = safe_str(row_vals.get(cols.get("actual_item"))) if cols.get("actual_item") else ""
                sticker_no = safe_int(row_vals.get(cols.get("sticker_no"))) if cols.get("sticker_no") else None
                bidder_name = safe_str(row_vals.get(cols.get("bidder_name"))) if cols.get("bidder_name") else ""
                bid_amount = safe_float(row_vals.get(cols.get("bid_amount"))) if cols.get("bid_amount") else 0
                fee = safe_float(row_vals.get(cols.get("fee"))) if cols.get("fee") else 0
                paid = safe_float(row_vals.get(cols.get("paid"))) if cols.get("paid") else 0
                handler = safe_str(row_vals.get(cols.get("handler"))) if cols.get("handler") else ""
                photo_codes = safe_str(row_vals.get(cols.get("photo_codes"))) if cols.get("photo_codes") else ""
                cost = safe_float(row_vals.get(cols.get("cost"))) if cols.get("cost") else 0
                source = safe_str(row_vals.get(cols.get("source"))) if cols.get("source") else ""

                # Skip Total rows
                if is_total_row(row_vals.get(cols.get("bidder_name"))):
                    continue

                # Skip rows with no meaningful data in mapped columns
                relevant_vals = {
                    "item_name": item_name,
                    "actual_item": actual_item,
                    "bidder_name": bidder_name,
                    "bid_amount": bid_amount,
                    "fee": fee,
                    "paid": paid,
                    "handler": handler,
                    "photo_codes": photo_codes,
                    "cost": cost,
                    "source": source,
                }
                if not any(v for v in relevant_vals.values()):
                    continue

                # ── Create ThisYearItem record ──
                if item_name or actual_item:
                    # Determine notes: combine handler, photo_codes, source as notes
                    notes_parts = []
                    if source:
                        notes_parts.append(f"來源:{source}")
                    notes_str = " | ".join(notes_parts)

                    this_year_item = ThisYearItem(
                        year=year,
                        sticker_no=sticker_no,
                        item_name=item_name,
                        actual_item=actual_item,
                        cost=cost if cost else 0,
                        source=source,
                        bidder_name=bidder_name,
                        bid_amount=bid_amount if bid_amount else 0,
                        paid_amount=paid if paid else 0,
                        handler=handler,
                        photo_codes=photo_codes,
                        notes=notes_str,
                    )
                    db.session.add(this_year_item)
                    year_stats["items"] += 1

                # ── Create Bid record (only if there's a named bidder) ──
                if bidder_name and bidder_name in member_registry:
                    member_id = member_registry[bidder_name]

                    # Look up item_id from registry
                    item_key = (item_name, actual_item)
                    item_id = item_registry.get(item_key)

                    bid = Bid(
                        year=year,
                        member_id=member_id,
                        item_id=item_id,
                        bid_amount=bid_amount if bid_amount else 0,
                        membership_fee=fee if fee else 0,
                        paid_amount=paid if paid else 0,
                        handler=handler,
                        photo_no=photo_codes,
                        bid_no=sticker_no,
                        source=source,
                    )
                    db.session.add(bid)
                    year_stats["bids"] += 1

                # ── Update Item.year_data with this year's info ──
                if item_key := (item_name, actual_item):
                    if item_key in item_registry:
                        item_id = item_registry[item_key]
                        item_obj = db.session.get(Item, item_id)
                        # Note: item_obj uses string PK (UUID), but item_id here is
                        # the business key. We need to look up by item_id field.
                        # We'll handle this below via a separate query.

            db.session.commit()
            total_this_year_items += year_stats["items"]
            total_bids += year_stats["bids"]
            stats[year] = year_stats
            print(f"  ✓ {year}: {year_stats['items']} items, {year_stats['bids']} bids")

        # ── 8. Update Item.year_data per-year ──
        # Collect all year data from this_year_items
        print()
        print("─── Building Item.year_data ───")
        items_data = {}  # (item_name, actual_item) -> list of {year, cost, source}

        items_all = ThisYearItem.query.all()
        for tyi in items_all:
            key = (tyi.item_name, tyi.actual_item)
            if key not in items_data:
                items_data[key] = []
            items_data[key].append({
                "year": tyi.year,
                "cost_hkd": tyi.cost,
                "source": tyi.source,
            })

        # Update Item records
        updated_items = 0
        for key, year_list in items_data.items():
            if key in item_registry:
                item_obj = Item.query.filter_by(item_id=item_registry[key]).first()
                if item_obj:
                    item_obj.year_data = json.dumps(year_list, ensure_ascii=False)
                    updated_items += 1
        db.session.commit()
        print(f"  ✓ Updated {updated_items} Item records with year data.")
        print()

        # ── Summary ──
        print("═" * 60)
        print("IMPORT SUMMARY")
        print("═" * 60)
        total_members = Member.query.count()
        total_items = Item.query.count()
        total_this_year = ThisYearItem.query.count()
        total_bid_count = Bid.query.count()
        print(f"  Members        : {total_members}")
        print(f"  Items (master) : {total_items}")
        print(f"  ThisYearItems  : {total_this_year_items}")
        print(f"  Bids           : {total_bid_count}")
        print()
        print("Per-year breakdown:")
        for year in sorted(stats.keys()):
            s = stats[year]
            print(f"  {year}: {s['items']} items, {s['bids']} bids")
        print()
        print("✓ Import complete!")


if __name__ == "__main__":
    main()
