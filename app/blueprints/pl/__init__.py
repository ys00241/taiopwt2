"""PL blueprint — 損益表 (Profit & Loss) 檢視與 Excel 匯出."""
from flask import Blueprint
from flask_login import login_required

bp = Blueprint("pl", __name__)


@bp.route("/pl")
@login_required
def view_pl():
    """View profit & loss entries by year with income/expense sections."""
    from flask import render_template, request
    from app.extensions import db
    from app.models.pl import PL
    from app.models.edition import Edition
    from app.models.this_year_item import ThisYearItem
    from app.models.live_income import LiveIncome
    from app.models.expense import Expense
    from app.models.daily_entry import DailyEntry
    from sqlalchemy import func
    from datetime import datetime

    year = request.args.get("year", type=int)

    # Default to latest year
    if not year or year == 0:
        latest = db.session.query(func.max(Edition.year)).scalar()
        year = latest or datetime.now().year

    # ════════════════════════════════════════════════════════════
    #  Bidding income — only THIS YEAR's items (no cross-year)
    # ════════════════════════════════════════════════════════════
    # 應收 = SUM(bid_amount) for this year's items
    bidding_receivable = (
        db.session.query(func.coalesce(func.sum(ThisYearItem.bid_amount), 0))
        .filter(ThisYearItem.year == year, ThisYearItem.bid_amount > 0)
        .scalar()
    ) or 0

    # 已收 = paid_amount on this year's items + live_income for this source_year
    bidding_paid_in_items = (
        db.session.query(func.coalesce(func.sum(ThisYearItem.paid_amount), 0))
        .filter(ThisYearItem.year == year, ThisYearItem.paid_amount > 0)
        .scalar()
    ) or 0

    bidding_live_collected = (
        db.session.query(func.coalesce(func.sum(LiveIncome.amount), 0))
        .filter(LiveIncome.source_year == year)
        .scalar()
    ) or 0

    bidding_received = bidding_paid_in_items + bidding_live_collected
    bidding_outstanding = bidding_receivable - bidding_received

    # ════════════════════════════════════════════════════════════
    #  PL table (static imported data — already received cash)
    # ════════════════════════════════════════════════════════════
    income_rows = (
        PL.query.filter(PL.year == year, PL.pl_type == "收入")
        .order_by(PL.subject)
        .all()
    )
    expense_rows = (
        PL.query.filter(PL.year == year, PL.pl_type == "支出")
        .order_by(PL.subject)
        .all()
    )

    # Build summary categorised by subject
    income_sections = {
        "收入(上年)": [],
        "收入(本年)": [],
        "會費": [],
        "香油": [],
        "其他收入": [],
    }
    for row in income_rows:
        subj = row.subject or ""
        if "上年" in subj or "上屆" in subj:
            income_sections["收入(上年)"].append(row)
        elif "本年" in subj or "本屆" in subj or "今屆" in subj:
            income_sections["收入(本年)"].append(row)
        elif "會費" in subj:
            income_sections["會費"].append(row)
        elif "香油" in subj or "香火" in subj:
            income_sections["香油"].append(row)
        else:
            income_sections["其他收入"].append(row)

    # ════════════════════════════════════════════════════════════
    #  DailyEntry records — 現金收支日記帳
    # ════════════════════════════════════════════════════════════
    daily_income_rows = (
        DailyEntry.query.filter(
            DailyEntry.year == year,
            DailyEntry.entry_type == "income",
        )
        .order_by(DailyEntry.entry_date.desc())
        .all()
    )
    daily_expense_rows = (
        DailyEntry.query.filter(
            DailyEntry.year == year,
            DailyEntry.entry_type == "expense",
        )
        .order_by(DailyEntry.entry_date.desc())
        .all()
    )

    # Normalise column name for Jinja2 template (DailyEntry uses 'amount', PL uses 'amount_hkd')
    for row in daily_income_rows:
        row.amount_hkd = row.amount
    for row in daily_expense_rows:
        row.amount_hkd = row.amount

    if daily_income_rows:
        income_sections["日常收入 (現金)"] = daily_income_rows

    income_totals = {}
    for section, rows in income_sections.items():
        income_totals[section] = sum(r.amount_hkd or 0 for r in rows)

    expense_sections = {}
    for row in expense_rows:
        cat = row.subject or "其他支出"
        if cat not in expense_sections:
            expense_sections[cat] = []
        expense_sections[cat].append(row)

    if daily_expense_rows:
        expense_sections["日常支出 (現金)"] = daily_expense_rows

    expense_totals = {}
    for section, rows in expense_sections.items():
        expense_totals[section] = sum(r.amount_hkd or 0 for r in rows)

    # Combined totals (PL + DailyEntry + live expenses)
    pl_total_income = sum(
        sum(r.amount_hkd or 0 for r in rows) for rows in income_sections.values()
    )
    total_expense = sum(
        sum(r.amount_hkd or 0 for r in rows) for rows in expense_sections.values()
    )

    # Also add live expenses (recorded on-site)
    live_expenses = (
        db.session.query(func.coalesce(func.sum(Expense.amount), 0))
        .filter(Expense.year == year, Expense.source == "live")
        .scalar()
    ) or 0
    total_expense += live_expenses

    # ════════════════════════════════════════════════════════════
    #  KEY METRICS
    # ════════════════════════════════════════════════════════════
    # 實收 = PL table income (already cash) + 競投已收
    actual_received = pl_total_income + bidding_received
    # 應收(名義) = PL table income + 全部競投應收額
    nominal_income = pl_total_income + bidding_receivable

    # 實收盈虧 = 實收 - 支出 ✅ REAL
    actual_surplus = actual_received - total_expense
    # 應收盈虧(名義) = 應收(名義) - 支出
    nominal_surplus = nominal_income - total_expense

    # ════════════════════════════════════════════════════════════
    #  Year list (from all data sources)
    # ════════════════════════════════════════════════════════════
    years_set = set()
    for r in db.session.query(func.distinct(PL.year)).all():
        if r[0]:
            years_set.add(r[0])
    for r in db.session.query(func.distinct(ThisYearItem.year)).all():
        if r[0]:
            years_set.add(r[0])
    for r in db.session.query(func.distinct(DailyEntry.year)).all():
        if r[0]:
            years_set.add(r[0])
    years = sorted(years_set, reverse=True)
    if not years:
        years = [datetime.now().year]
    if year not in years:
        year = years[0]

    # ════════════════════════════════════════════════════════════
    #  Multi-year comparison data
    # ════════════════════════════════════════════════════════════
    multi_year_data = {}
    for y in years:
        pl_inc = (
            db.session.query(func.coalesce(func.sum(PL.amount_hkd), 0))
            .filter(PL.year == y, PL.pl_type == "收入")
            .scalar()
        ) or 0
        pl_exp = (
            db.session.query(func.coalesce(func.sum(PL.amount_hkd), 0))
            .filter(PL.year == y, PL.pl_type == "支出")
            .scalar()
        ) or 0
        live_exp = (
            db.session.query(func.coalesce(func.sum(Expense.amount), 0))
            .filter(Expense.year == y, Expense.source == "live")
            .scalar()
        ) or 0
        daily_inc = (
            db.session.query(func.coalesce(func.sum(DailyEntry.amount), 0))
            .filter(DailyEntry.year == y, DailyEntry.entry_type == "income")
            .scalar()
        ) or 0
        daily_exp = (
            db.session.query(func.coalesce(func.sum(DailyEntry.amount), 0))
            .filter(DailyEntry.year == y, DailyEntry.entry_type == "expense")
            .scalar()
        ) or 0
        bid_amt = (
            db.session.query(func.coalesce(func.sum(ThisYearItem.bid_amount), 0))
            .filter(ThisYearItem.year == y, ThisYearItem.bid_amount > 0)
            .scalar()
        ) or 0
        paid_in_items = (
            db.session.query(func.coalesce(func.sum(ThisYearItem.paid_amount), 0))
            .filter(ThisYearItem.year == y, ThisYearItem.paid_amount > 0)
            .scalar()
        ) or 0
        live_collected = (
            db.session.query(func.coalesce(func.sum(LiveIncome.amount), 0))
            .filter(LiveIncome.source_year == y)
            .scalar()
        ) or 0

        total_inc = pl_inc + daily_inc
        bid_rcvd = paid_in_items + live_collected
        actual_rcvd = total_inc + bid_rcvd
        total_exp = pl_exp + live_exp + daily_exp

        multi_year_data[y] = {
            "pl_total_income": total_inc,
            "bidding_receivable": bid_amt,
            "actual_received": actual_rcvd,
            "total_expense": total_exp,
            "actual_surplus": actual_rcvd - total_exp,
        }

    return render_template(
        "pl/index.html",
        year=year,
        years=years,
        income_sections=income_sections,
        income_totals=income_totals,
        expense_sections=expense_sections,
        expense_totals=expense_totals,
        total_expense=total_expense,
        pl_total_income=pl_total_income,
        # Bidding breakdown
        bidding_receivable=bidding_receivable,
        bidding_received=bidding_received,
        bidding_outstanding=bidding_outstanding,
        # Real vs nominal
        actual_received=actual_received,
        nominal_income=nominal_income,
        actual_surplus=actual_surplus,
        nominal_surplus=nominal_surplus,
        # Multi-year comparison
        multi_year_data=multi_year_data,
    )


@bp.route("/pl/ledger-data")
@login_required
def pl_ledger_data():
    """Return PL ledger data as JSON — combined PL + DailyEntry rows for a year."""
    from flask import jsonify, request
    from app.extensions import db
    from app.models.pl import PL
    from app.models.daily_entry import DailyEntry

    year = request.args.get("year", type=int)
    if not year:
        return jsonify({"error": "year parameter required"}), 400

    # Query PL rows
    pl_rows = PL.query.filter(PL.year == year).order_by(PL.subject).all()
    # Query DailyEntry rows
    de_rows = DailyEntry.query.filter(DailyEntry.year == year).order_by(DailyEntry.entry_date).all()

    ledger = []

    for r in pl_rows:
        ledger.append({
            "date": str(r.year),
            "type": r.pl_type or "",
            "subject": r.subject or "",
            "amount": r.amount_hkd or 0,
            "payment_method": r.payment_method or "",
            "notes": "",
            "source": "PL",
        })

    for r in de_rows:
        entry_type = "收入" if r.entry_type == "income" else "支出"
        ledger.append({
            "date": str(r.entry_date) if r.entry_date else str(r.year),
            "type": entry_type,
            "subject": r.subject or "",
            "amount": r.amount or 0,
            "payment_method": r.payment_method or "",
            "notes": r.notes or "",
            "source": "DailyEntry",
        })

    # Sort by date
    ledger.sort(key=lambda x: x["date"])

    total_income = sum(item["amount"] for item in ledger if item["type"] == "收入")
    total_expense = sum(item["amount"] for item in ledger if item["type"] == "支出")

    return jsonify({
        "ledger": ledger,
        "total_income": total_income,
        "total_expense": total_expense,
    })


@bp.route("/pl/export")
@login_required
def export_pl_excel():
    """Export P&L as Excel (.xlsx) with multi-year comparison."""
    from flask import request, Response
    from app.extensions import db
    from app.models.pl import PL
    from app.models.this_year_item import ThisYearItem
    from app.models.live_income import LiveIncome
    from app.models.expense import Expense
    from app.models.daily_entry import DailyEntry
    from app.models.edition import Edition
    from sqlalchemy import func
    import io

    years_param = request.args.get("years", "")
    if years_param:
        selected_years = [int(y.strip()) for y in years_param.split(",") if y.strip()]
    else:
        selected_years = [
            r[0]
            for r in db.session.query(func.distinct(PL.year))
            .order_by(PL.year.desc())
            .limit(3)
            .all()
        ]
        if not selected_years:
            selected_years = [
                r[0]
                for r in db.session.query(func.distinct(ThisYearItem.year))
                .order_by(ThisYearItem.year.desc())
                .limit(3)
                .all()
            ]

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "損益表"

    title_font = Font(name="Microsoft JhengHei", size=14, bold=True)
    header_font = Font(name="Microsoft JhengHei", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    section_font = Font(name="Microsoft JhengHei", size=11, bold=True, color="1F4E79")
    section_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    data_font = Font(name="Microsoft JhengHei", size=10)
    total_font = Font(name="Microsoft JhengHei", size=11, bold=True)
    total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    money_fmt = '#,##0'

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + len(selected_years))
    title_cell = ws.cell(row=1, column=1, value="寶榮堂花炮會 — 損益表 (Profit & Loss)")
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center")

    headers = ["科目"] + [f"{y}年" for y in selected_years]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    def get_subject_totals(pl_type):
        rows = PL.query.filter(
            PL.pl_type == pl_type,
            PL.year.in_(selected_years),
        ).all()
        result = {}
        for r in rows:
            if r.subject not in result:
                result[r.subject] = {y: 0 for y in selected_years}
            result[r.subject][r.year] = (result[r.subject].get(r.year, 0) + (r.amount_hkd or 0))
        return result

    income_data = get_subject_totals("收入")
    expense_data = get_subject_totals("支出")

    # Bidding + live income per year
    bidding_income_per_year = {}
    bidding_received_per_year = {}
    for y in selected_years:
        bid_amt = (
            db.session.query(func.coalesce(func.sum(ThisYearItem.bid_amount), 0))
            .filter(ThisYearItem.year == y, ThisYearItem.bid_amount > 0)
            .scalar()
        ) or 0
        paid_in_items = (
            db.session.query(func.coalesce(func.sum(ThisYearItem.paid_amount), 0))
            .filter(ThisYearItem.year == y, ThisYearItem.paid_amount > 0)
            .scalar()
        ) or 0
        live_collected = (
            db.session.query(func.coalesce(func.sum(LiveIncome.amount), 0))
            .filter(LiveIncome.source_year == y)
            .scalar()
        ) or 0
        bidding_income_per_year[y] = bid_amt
        bidding_received_per_year[y] = paid_in_items + live_collected

    # Live expenses per year
    live_exp_per_year = {}
    for y in selected_years:
        le = (
            db.session.query(func.coalesce(func.sum(Expense.amount), 0))
            .filter(Expense.year == y, Expense.source == "live")
            .scalar()
        ) or 0
        live_exp_per_year[y] = le

    row_num = 4

    # ── Income Section ──
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=1 + len(selected_years))
    cell = ws.cell(row=row_num, column=1, value="收入")
    cell.font = section_font
    cell.fill = section_fill
    cell.border = thin_border
    for c in range(2, 2 + len(selected_years)):
        ws.cell(row=row_num, column=c).fill = section_fill
        ws.cell(row=row_num, column=c).border = thin_border
    row_num += 1

    for subject, yearly in sorted(income_data.items()):
        ws.cell(row=row_num, column=1, value=subject).font = data_font
        ws.cell(row=row_num, column=1).border = thin_border
        for ci, y in enumerate(selected_years, 2):
            cell = ws.cell(row=row_num, column=ci, value=yearly.get(y, 0) or 0)
            cell.font = data_font
            cell.number_format = money_fmt
            cell.border = thin_border
        row_num += 1

    # PL income total
    pl_year_totals = {}
    for subject, yearly in income_data.items():
        for y in selected_years:
            pl_year_totals[y] = pl_year_totals.get(y, 0) + (yearly.get(y, 0) or 0)

    ws.cell(row=row_num, column=1, value="PL收入合計 (已收)").font = data_font
    ws.cell(row=row_num, column=1).border = thin_border
    for ci, y in enumerate(selected_years, 2):
        cell = ws.cell(row=row_num, column=ci, value=pl_year_totals.get(y, 0))
        cell.font = data_font
        cell.number_format = money_fmt
        cell.border = thin_border
    row_num += 1

    # Bidding income row (應收)
    from copy import copy
    from openpyxl.styles import Font as OxlFont
    cell = ws.cell(row=row_num, column=1, value="競投應收 (即時)")
    cell.font = Font(name="Microsoft JhengHei", size=10, italic=True)
    cell.border = thin_border
    for ci, y in enumerate(selected_years, 2):
        cell = ws.cell(row=row_num, column=ci, value=bidding_income_per_year.get(y, 0))
        cell.font = data_font
        cell.number_format = money_fmt
        cell.border = thin_border
    row_num += 1

    # Bidding already received
    ws.cell(row=row_num, column=1, value="  已收競投").font = data_font
    ws.cell(row=row_num, column=1).border = thin_border
    for ci, y in enumerate(selected_years, 2):
        cell = ws.cell(row=row_num, column=ci, value=bidding_received_per_year.get(y, 0))
        cell.font = data_font
        cell.number_format = money_fmt
        cell.border = thin_border
    row_num += 1

    # Actual total received
    actual_received_totals = {}
    for y in selected_years:
        actual_received_totals[y] = pl_year_totals.get(y, 0) + bidding_received_per_year.get(y, 0)

    ws.cell(row=row_num, column=1, value="實收合計 (PL已收 + 競投已收)").font = total_font
    ws.cell(row=row_num, column=1).fill = total_fill
    ws.cell(row=row_num, column=1).border = thin_border
    for ci, y in enumerate(selected_years, 2):
        cell = ws.cell(row=row_num, column=ci, value=actual_received_totals.get(y, 0))
        cell.font = total_font
        cell.fill = total_fill
        cell.number_format = money_fmt
        cell.border = thin_border
    row_num += 1

    # Nominal total (PL + all bidding receivable)
    nominal_totals = {}
    for y in selected_years:
        nominal_totals[y] = pl_year_totals.get(y, 0) + bidding_income_per_year.get(y, 0)

    ws.cell(row=row_num, column=1, value="應收總計 (名義: PL + 競投應收)").font = total_font
    ws.cell(row=row_num, column=1).border = thin_border
    for ci, y in enumerate(selected_years, 2):
        cell = ws.cell(row=row_num, column=ci, value=nominal_totals.get(y, 0))
        cell.font = total_font
        cell.number_format = money_fmt
        cell.border = thin_border
    row_num += 2

    # ── Expense Section ──
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=1 + len(selected_years))
    cell = ws.cell(row=row_num, column=1, value="支出")
    cell.font = section_font
    cell.fill = section_fill
    cell.border = thin_border
    for c in range(2, 2 + len(selected_years)):
        ws.cell(row=row_num, column=c).fill = section_fill
        ws.cell(row=row_num, column=c).border = thin_border
    row_num += 1

    for subject, yearly in sorted(expense_data.items()):
        ws.cell(row=row_num, column=1, value=subject).font = data_font
        ws.cell(row=row_num, column=1).border = thin_border
        for ci, y in enumerate(selected_years, 2):
            cell = ws.cell(row=row_num, column=ci, value=yearly.get(y, 0) or 0)
            cell.font = data_font
            cell.number_format = money_fmt
            cell.border = thin_border
        row_num += 1

    # Live expenses row
    ws.cell(row=row_num, column=1, value="現場支出 (Live)").font = data_font
    ws.cell(row=row_num, column=1).border = thin_border
    for ci, y in enumerate(selected_years, 2):
        cell = ws.cell(row=row_num, column=ci, value=live_exp_per_year.get(y, 0))
        cell.font = data_font
        cell.number_format = money_fmt
        cell.border = thin_border
    row_num += 1

    # Total expense (PL + live)
    expense_year_totals = {}
    for subject, yearly in expense_data.items():
        for y in selected_years:
            expense_year_totals[y] = expense_year_totals.get(y, 0) + (yearly.get(y, 0) or 0)
    for y in selected_years:
        expense_year_totals[y] = expense_year_totals.get(y, 0) + live_exp_per_year.get(y, 0)

    ws.cell(row=row_num, column=1, value="支出合計").font = total_font
    ws.cell(row=row_num, column=1).fill = total_fill
    ws.cell(row=row_num, column=1).border = thin_border
    for ci, y in enumerate(selected_years, 2):
        cell = ws.cell(row=row_num, column=ci, value=expense_year_totals.get(y, 0))
        cell.font = total_font
        cell.fill = total_fill
        cell.number_format = money_fmt
        cell.border = thin_border
    row_num += 1

    # ── Surplus ──
    # Actual surplus
    ws.cell(row=row_num, column=1, value="實收盈虧 (實收 - 支出)").font = total_font
    ws.cell(row=row_num, column=1).fill = total_fill
    ws.cell(row=row_num, column=1).border = thin_border
    for ci, y in enumerate(selected_years, 2):
        net = actual_received_totals.get(y, 0) - expense_year_totals.get(y, 0)
        cell = ws.cell(row=row_num, column=ci, value=net)
        cell.font = total_font
        cell.fill = total_fill
        cell.number_format = money_fmt
        cell.border = thin_border
    row_num += 1

    # Nominal surplus
    ws.cell(row=row_num, column=1, value="應收盈虧 (名義: 應收 - 支出)").font = total_font
    ws.cell(row=row_num, column=1).border = thin_border
    for ci, y in enumerate(selected_years, 2):
        net = nominal_totals.get(y, 0) - expense_year_totals.get(y, 0)
        cell = ws.cell(row=row_num, column=ci, value=net)
        cell.font = total_font
        cell.number_format = money_fmt
        cell.border = thin_border
    row_num += 1

    ws.column_dimensions["A"].width = 35
    for ci in range(2, 2 + len(selected_years)):
        ws.column_dimensions[chr(64 + ci)].width = 18

    # ── Sheet 2: 細則 Ledger ──
    ws2 = wb.create_sheet("細則 Ledger")
    ledger_headers = ["日期", "類型", "科目", "金額", "付款方式", "備註", "來源"]
    for ci, h in enumerate(ledger_headers, 1):
        cell = ws2.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Query ALL PL rows for selected years
    all_pl = (
        PL.query
        .filter(PL.year.in_(selected_years))
        .order_by(PL.year, PL.subject)
        .all()
    )
    # Query DailyEntry rows for selected years
    all_de = (
        DailyEntry.query
        .filter(DailyEntry.year.in_(selected_years))
        .order_by(DailyEntry.year, DailyEntry.entry_date)
        .all()
    )

    ledger_rows = []
    for r in all_pl:
        ledger_rows.append({
            "date": str(r.year),
            "type": r.pl_type or "",
            "subject": r.subject or "",
            "amount": r.amount_hkd or 0,
            "payment_method": r.payment_method or "",
            "notes": "",
            "source": "PL",
        })
    for r in all_de:
        entry_type = "收入" if r.entry_type == "income" else "支出"
        ledger_rows.append({
            "date": str(r.entry_date) if r.entry_date else str(r.year),
            "type": entry_type,
            "subject": r.subject or "",
            "amount": r.amount or 0,
            "payment_method": r.payment_method or "",
            "notes": r.notes or "",
            "source": "日記帳",
        })

    ledger_rows.sort(key=lambda x: x["date"])

    for ri, lr in enumerate(ledger_rows, 2):
        ws2.cell(row=ri, column=1, value=lr["date"]).font = data_font
        ws2.cell(row=ri, column=1).border = thin_border
        ws2.cell(row=ri, column=2, value=lr["type"]).font = data_font
        ws2.cell(row=ri, column=2).border = thin_border
        ws2.cell(row=ri, column=3, value=lr["subject"]).font = data_font
        ws2.cell(row=ri, column=3).border = thin_border
        cell = ws2.cell(row=ri, column=4, value=lr["amount"])
        cell.font = data_font
        cell.number_format = money_fmt
        cell.border = thin_border
        ws2.cell(row=ri, column=5, value=lr["payment_method"]).font = data_font
        ws2.cell(row=ri, column=5).border = thin_border
        ws2.cell(row=ri, column=6, value=lr["notes"]).font = data_font
        ws2.cell(row=ri, column=6).border = thin_border
        ws2.cell(row=ri, column=7, value=lr["source"]).font = data_font
        ws2.cell(row=ri, column=7).border = thin_border

    # Auto-fit column widths for ledger sheet
    col_widths = [14, 8, 30, 14, 14, 30, 12]
    for ci, w in enumerate(col_widths, 1):
        ws2.column_dimensions[chr(64 + ci)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    years_label = "_".join(str(y) for y in selected_years)
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="pl_{years_label}.xlsx"',
        },
    )
